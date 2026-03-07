#!/usr/bin/env python3
"""
Deployment/Retrieval Validation Workflow

Two modes:
  1. GENERATE: Run deployment detection on videos, produce an XLSX for review.
     python validate_deployment.py generate -i /path/to/videos -o deployment_review.xlsx

  2. APPLY: Ingest the reviewed XLSX, override auto-detected times with user
     corrections, and produce a final CSV used by the pipeline.
     python validate_deployment.py apply -i deployment_review.xlsx -o stable_durations.csv

The XLSX has columns:
  video_file | duration | deploy_auto (HH:MM:SS) | deploy_validated (HH:MM:SS) |
  retrieve_auto (HH:MM:SS) | retrieve_validated (HH:MM:SS) | stable_duration | notes

Users fill in deploy_validated / retrieve_validated only where corrections are needed.
Blank cells mean the auto-detected value is accepted.
"""

import argparse
import os
import sys
from pathlib import Path
from datetime import timedelta


def _sec_to_hms(seconds):
    """Convert seconds (float) to HH:MM:SS string."""
    if seconds is None or seconds != seconds:  # NaN check
        return ""
    td = timedelta(seconds=round(seconds))
    total_sec = int(td.total_seconds())
    h = total_sec // 3600
    m = (total_sec % 3600) // 60
    s = total_sec % 60
    return f"{h:02d}:{m:02d}:{s:02d}"


def _hms_to_sec(value):
    """Convert HH:MM:SS (string or datetime.time) to seconds. Returns None if blank/invalid."""
    if value is None:
        return None

    # openpyxl may return a datetime.time object for time-formatted cells
    import datetime
    if isinstance(value, datetime.time):
        return value.hour * 3600 + value.minute * 60 + value.second

    if isinstance(value, (int, float)):
        # Excel stores times as fractions of a day
        if value < 1:  # fraction of day
            return round(value * 86400)
        return round(value)

    s = str(value).strip()
    if not s:
        return None

    parts = s.split(":")
    try:
        if len(parts) == 3:
            return int(parts[0]) * 3600 + int(parts[1]) * 60 + int(parts[2])
        elif len(parts) == 2:
            return int(parts[0]) * 60 + int(parts[1])
        else:
            return round(float(s))
    except (ValueError, TypeError):
        return None


def generate(input_path, output_xlsx, recursive=True):
    """Run deployment detection on all videos and write review XLSX."""
    from utils.deployment_detector import DeploymentDetector
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, numbers

    input_path = Path(input_path)

    # Find videos
    if input_path.is_file():
        videos = [input_path]
    else:
        pattern = "**/*.MP4" if recursive else "*.MP4"
        videos = sorted(input_path.glob(pattern))
        if not videos:
            videos = sorted(input_path.glob(pattern.replace(".MP4", ".mp4")))
        if not videos:
            # Try AVI as well
            videos = sorted(input_path.glob(pattern.replace(".MP4", ".avi")))

    if not videos:
        print(f"No video files found in {input_path}")
        sys.exit(1)

    print(f"Found {len(videos)} videos. Running deployment detection...")

    detector = DeploymentDetector()
    rows = []

    for i, vp in enumerate(videos):
        print(f"\n[{i+1}/{len(videos)}] {vp.name}")
        try:
            times = detector.get_stable_times(str(vp))
            rows.append({
                'video_file': vp.name,
                'video_path': str(vp),
                'duration_sec': times['total_duration_sec'],
                'deploy_sec': times['deploy_sec'],
                'retrieve_end_sec': times['stable_end_sec'],
                'retrieve_sec': times['retrieve_sec'],
            })
        except Exception as e:
            print(f"  Error: {e}")
            rows.append({
                'video_file': vp.name,
                'video_path': str(vp),
                'duration_sec': 0,
                'deploy_sec': 0,
                'retrieve_end_sec': 0,
                'retrieve_sec': 0,
            })

    # Build XLSX
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Deployment Review"

    # Headers
    headers = [
        ("A", "video_file", 40),
        ("B", "duration", 12),
        ("C", "deploy_auto", 14),
        ("D", "deploy_validated", 18),
        ("E", "retrieve_auto", 14),
        ("F", "retrieve_validated", 18),
        ("G", "stable_duration", 16),
        ("H", "notes", 30),
    ]

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    editable_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    time_format = "HH:MM:SS"

    for col_letter, name, width in headers:
        cell = ws[f"{col_letter}1"]
        cell.value = name
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[col_letter].width = width

    # Data rows
    for idx, row in enumerate(rows, start=2):
        deploy_sec = row['deploy_sec']
        retrieve_end = row['retrieve_end_sec']
        total_sec = row['duration_sec']

        # A: video filename
        ws[f"A{idx}"] = row['video_file']

        # B: total duration as HH:MM:SS
        ws[f"B{idx}"] = _sec_to_hms(total_sec)
        ws[f"B{idx}"].alignment = Alignment(horizontal="center")

        # C: auto-detected deployment time (HH:MM:SS)
        ws[f"C{idx}"] = _sec_to_hms(deploy_sec)
        ws[f"C{idx}"].alignment = Alignment(horizontal="center")

        # D: user-validated deployment (editable, yellow background)
        ws[f"D{idx}"].fill = editable_fill
        ws[f"D{idx}"].number_format = time_format
        ws[f"D{idx}"].alignment = Alignment(horizontal="center")

        # E: auto-detected retrieval time (time in video, not duration of retrieval)
        ws[f"E{idx}"] = _sec_to_hms(retrieve_end)
        ws[f"E{idx}"].alignment = Alignment(horizontal="center")

        # F: user-validated retrieval (editable, yellow background)
        ws[f"F{idx}"].fill = editable_fill
        ws[f"F{idx}"].number_format = time_format
        ws[f"F{idx}"].alignment = Alignment(horizontal="center")

        # G: stable duration formula (auto-calculated)
        # stable = (validated_retrieve or auto_retrieve) - (validated_deploy or auto_deploy)
        # We store as text for simplicity — recalculated at apply time
        stable_sec = retrieve_end - deploy_sec
        ws[f"G{idx}"] = _sec_to_hms(max(0, stable_sec))
        ws[f"G{idx}"].alignment = Alignment(horizontal="center")

        # H: notes (editable, yellow background)
        ws[f"H{idx}"].fill = editable_fill

    # Hidden column I: full video path (for apply step)
    ws.column_dimensions["I"].hidden = True
    ws["I1"] = "video_path"
    ws["I1"].font = header_font
    ws["I1"].fill = header_fill
    for idx, row in enumerate(rows, start=2):
        ws[f"I{idx}"] = row['video_path']

    # Instructions sheet
    instr = wb.create_sheet("Instructions")
    instructions = [
        "DEPLOYMENT/RETRIEVAL VALIDATION",
        "",
        "Yellow cells are editable. All other cells are auto-generated.",
        "",
        "COLUMNS:",
        "  video_file       - Video filename",
        "  duration          - Total video duration",
        "  deploy_auto       - Auto-detected deployment end time (stable period starts here)",
        "  deploy_validated  - YOUR correction if auto is wrong (HH:MM:SS). Leave blank to accept auto.",
        "  retrieve_auto     - Auto-detected retrieval start time (stable period ends here)",
        "  retrieve_validated- YOUR correction if auto is wrong (HH:MM:SS). Leave blank to accept auto.",
        "  stable_duration   - Calculated stable recording duration",
        "  notes             - Any notes about this video",
        "",
        "HOW TO VALIDATE:",
        "  1. Open each video in VLC or similar",
        "  2. Find the moment the camera settles on the seabed (deployment end)",
        "  3. Find the moment retrieval begins (camera lifts off seabed)",
        "  4. Enter corrected times in the yellow cells ONLY where the auto value is wrong",
        "  5. Use HH:MM:SS format (e.g. 00:01:30 for 1 minute 30 seconds)",
        "  6. Save the file and run: python validate_deployment.py apply -i this_file.xlsx",
        "",
        "TIPS:",
        "  - For videos where the camera never stabilises, set deploy_validated = duration",
        "  - For videos where retrieval is not captured (battery died), leave retrieve_validated blank",
        "  - The 'notes' column is for your reference only; it is preserved but not used by the pipeline",
    ]
    for i, line in enumerate(instructions, start=1):
        instr[f"A{i}"] = line
    instr.column_dimensions["A"].width = 100

    # Freeze header row
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:H{len(rows)+1}"

    wb.save(output_xlsx)
    print(f"\nSaved review spreadsheet: {output_xlsx}")
    print(f"  {len(rows)} videos")
    print(f"  Edit yellow cells, then run:")
    print(f"  python validate_deployment.py apply -i {output_xlsx}")


def apply(input_xlsx, output_csv):
    """Read reviewed XLSX, merge user corrections with auto values, write final CSV."""
    import openpyxl
    import csv

    wb = openpyxl.load_workbook(input_xlsx, data_only=True)
    ws = wb["Deployment Review"]

    rows_out = []
    for row_idx in range(2, ws.max_row + 1):
        video_file = ws[f"A{row_idx}"].value
        if not video_file:
            continue

        video_path = ws[f"I{row_idx}"].value or ""
        duration_str = ws[f"B{row_idx}"].value
        deploy_auto_str = ws[f"C{row_idx}"].value
        deploy_valid_val = ws[f"D{row_idx}"].value
        retrieve_auto_str = ws[f"E{row_idx}"].value
        retrieve_valid_val = ws[f"F{row_idx}"].value
        notes = ws[f"H{row_idx}"].value or ""

        # Parse times
        duration_sec = _hms_to_sec(duration_str) or 0
        deploy_auto_sec = _hms_to_sec(deploy_auto_str) or 0
        retrieve_auto_sec = _hms_to_sec(retrieve_auto_str) or duration_sec

        deploy_valid_sec = _hms_to_sec(deploy_valid_val)
        retrieve_valid_sec = _hms_to_sec(retrieve_valid_val)

        # Use validated value if provided, otherwise auto
        deploy_final = deploy_valid_sec if deploy_valid_sec is not None else deploy_auto_sec
        retrieve_final = retrieve_valid_sec if retrieve_valid_sec is not None else retrieve_auto_sec

        stable_sec = max(0, retrieve_final - deploy_final)

        rows_out.append({
            'video_file': video_file,
            'video_path': video_path,
            'total_duration_sec': duration_sec,
            'deploy_auto_sec': deploy_auto_sec,
            'deploy_validated_sec': deploy_valid_sec if deploy_valid_sec is not None else "",
            'deploy_final_sec': deploy_final,
            'retrieve_auto_sec': retrieve_auto_sec,
            'retrieve_validated_sec': retrieve_valid_sec if retrieve_valid_sec is not None else "",
            'retrieve_final_sec': retrieve_final,
            'stable_duration_sec': stable_sec,
            'stable_duration_min': round(stable_sec / 60, 1),
            'notes': notes,
        })

    # Write CSV
    if not rows_out:
        print("No data rows found in spreadsheet.")
        sys.exit(1)

    fieldnames = list(rows_out[0].keys())
    with open(output_csv, 'w', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows_out)

    # Summary
    n_corrected_deploy = sum(1 for r in rows_out if r['deploy_validated_sec'] != "")
    n_corrected_retrieve = sum(1 for r in rows_out if r['retrieve_validated_sec'] != "")

    print(f"\nSaved: {output_csv}")
    print(f"  {len(rows_out)} videos")
    print(f"  {n_corrected_deploy} deployment corrections applied")
    print(f"  {n_corrected_retrieve} retrieval corrections applied")
    print(f"  {len(rows_out) - n_corrected_deploy} deployments accepted as auto-detected")
    print(f"  {len(rows_out) - n_corrected_retrieve} retrievals accepted as auto-detected")


def main():
    parser = argparse.ArgumentParser(
        description="Deployment/Retrieval Validation Workflow",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Step 1: Generate review spreadsheet
  python validate_deployment.py generate -i /path/to/videos -o deployment_review.xlsx

  # Step 2: Edit deployment_review.xlsx in Excel/LibreOffice (fill yellow cells)

  # Step 3: Apply corrections
  python validate_deployment.py apply -i deployment_review.xlsx -o stable_durations.csv
        """
    )

    subparsers = parser.add_subparsers(dest="command", required=True)

    # Generate subcommand
    gen = subparsers.add_parser("generate", help="Run detection and create review XLSX")
    gen.add_argument("-i", "--input", required=True, help="Video file or directory")
    gen.add_argument("-o", "--output", default="deployment_review.xlsx", help="Output XLSX path")
    gen.add_argument("--no-recursive", action="store_true", help="Don't search subdirectories")

    # Apply subcommand
    app = subparsers.add_parser("apply", help="Ingest reviewed XLSX and produce final CSV")
    app.add_argument("-i", "--input", required=True, help="Reviewed XLSX file")
    app.add_argument("-o", "--output", default="stable_durations.csv", help="Output CSV path")

    args = parser.parse_args()

    if args.command == "generate":
        generate(args.input, args.output, recursive=not args.no_recursive)
    elif args.command == "apply":
        apply(args.input, args.output)


if __name__ == "__main__":
    main()
